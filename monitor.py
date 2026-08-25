import os
import sys
import time
import json
import re
import traceback
from datetime import datetime
import pandas as pd
import openpyxl
from playwright.sync_api import sync_playwright

# Configuration
CHECK_INTERVAL_SECONDS = 3600  # Check every hour (3600 seconds)
STATE_FILE = "state.json"
URL = "https://csms.bpcl.statiq.co.in/charger-management/all-chargers"
EXCEL_PATH = r"E:\BPCL office\EVCS Access\Chargers Master.xlsx"
if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = "Chargers Master.xlsx"

UPTIME_DIR = r"E:\BPCL office\EVCS Access\Uptime"
if not os.path.exists(UPTIME_DIR):
    UPTIME_DIR = "Uptime"

DASHBOARD_HTML_PATH = "dashboard.html"

def log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")

def parse_status_from_cell(cell, page):
    """
    Detects if the status symbol is online (green) or offline (red)
    by checking image src attributes, classes, text, and computed style colors.
    """
    # 1. Check for images (e.g. disconnected.svg vs connected.svg)
    images = cell.query_selector_all("img")
    for img in images:
        src = img.get_attribute("src") or ""
        src = src.lower()
        if "disconnected" in src or "offline" in src:
            return "Offline"
        if "connected" in src or "available" in src or "online" in src or "charging" in src:
            return "Online"
            
    # 2. Check for standard text clues
    text = cell.inner_text().strip().lower()
    if "online" in text:
        return "Online"
    if "offline" in text:
        return "Offline"
    
    # 3. Check all elements inside cell for green/red indicators
    elements = cell.query_selector_all("svg, img, span, div, i")
    for el in elements:
        try:
            class_name = el.get_attribute("class") or ""
            class_name = class_name.lower()
            if "online" in class_name or "success" in class_name:
                return "Online"
            if "offline" in class_name or "danger" in class_name or "error" in class_name:
                return "Offline"
            
            # Fetch computed styles from browser
            styles = page.evaluate("""el => {
                const s = window.getComputedStyle(el);
                return {
                    color: s.color,
                    backgroundColor: s.backgroundColor,
                    fill: s.fill,
                    stroke: s.stroke
                };
            }""", el)
            
            # Check colors
            for prop in ["fill", "stroke", "color", "backgroundColor"]:
                val = styles.get(prop, "")
                if not val or val == "none" or val == "rgba(0, 0, 0, 0)":
                    continue
                # Simple color string matches
                if "rgb(16, 185, 129)" in val or "rgb(74, 222, 128)" in val or "green" in val:
                    return "Online"
                if "rgb(239, 68, 68)" in val or "rgb(248, 113, 113)" in val or "red" in val:
                    return "Offline"
        except Exception:
            continue
            
    return "Offline"  # Default fallback if status is unclear

def click_next_page_if_available(page):
    """
    Finds the next page button and clicks it if available.
    """
    try:
        clicked_info = page.evaluate("""() => {
            const elements = Array.from(document.querySelectorAll('button, a, div, span, li, svg, i'));
            const ofElement = elements.find(el => {
                if (el.children.length > 0) return false;
                const text = el.innerText || el.textContent || "";
                return text.trim().startsWith("of") && /\\d+/.test(text);
            });
            
            if (ofElement) {
                const clickables = elements.filter(el => {
                    const isVisible = el.offsetWidth > 0 && el.offsetHeight > 0;
                    if (!isVisible) return false;
                    const tag = el.tagName.toLowerCase();
                    const isClickableTag = tag === 'button' || tag === 'a' || tag === 'svg' || tag === 'i';
                    return isClickableTag && !el.contains(ofElement);
                });
                
                const elementsAfter = clickables.filter(el => {
                    return (ofElement.compareDocumentPosition(el) & Node.DOCUMENT_POSITION_FOLLOWING) !== 0;
                });
                
                if (elementsAfter.length >= 1) {
                    const nextBtn = elementsAfter[0];
                    const isDisabled = nextBtn.hasAttribute('disabled') || 
                                       nextBtn.getAttribute('aria-disabled') === 'true' ||
                                       nextBtn.className.includes('disabled');
                    if (!isDisabled) {
                        nextBtn.click();
                        return "Clicked next button";
                    }
                }
            }
            return "";
        }""")
        return bool(clicked_info)
    except Exception as e:
        log(f"[WARNING] Error in click_next_page_if_available: {e}")
        return False

def change_rows_per_page_to_100(page):
    """
    Attempts to find the 'Rows per page' dropdown and set it to 100.
    """
    try:
        log("Attempting to change rows per page to 100...")
        page.wait_for_selector("table tbody tr", timeout=15000)
        
        combobox = page.locator(".pagination_container div[role='combobox']")
        if combobox.count() > 0:
            text_val = combobox.inner_text().strip().replace('\u200b', '')
            if text_val == "100":
                log("Rows per page is already set to 100.")
                return
                
            log(f"Current rows per page value: {text_val}. Clicking dropdown...")
            combobox.click()
            page.wait_for_timeout(1500)
            
            option_100 = page.locator("li[role='option']:has-text('100')")
            if option_100.count() == 0:
                option_100 = page.locator("li:has-text('100'), [role='option']:has-text('100')").first
                
            if option_100.count() > 0:
                option_100.click()
                log("Clicked 100 rows per page. Waiting for table reload...")
                page.wait_for_timeout(5000)
            else:
                log("[WARNING] Option 100 not found in dropdown.")
        else:
            log("[WARNING] Pagination combobox not found on page.")
    except Exception as e:
        log(f"[WARNING] Could not change rows per page to 100: {e}")

def scrape_chargers(page):
    """Scrapes all chargers and returns a list of dicts."""
    all_chargers = []
    
    # Wait for table to load
    try:
        page.wait_for_selector("table", timeout=15000)
    except Exception:
        log("[ERROR] Table element not found on page.")
        return []
        
    table = page.query_selector("table")
    if not table:
        return []
        
    # Get column headers
    header_cells = table.query_selector_all("thead th, tr th")
    headers = [cell.inner_text().strip() for cell in header_cells]
    if not headers:
        log("[ERROR] Could not extract table headers.")
        return []
        
    header_map = {name: idx for idx, name in enumerate(headers)}
    
    # Get rows
    data_rows = table.query_selector_all("tbody tr")
    log(f"Found {len(data_rows)} rows on the current page. Parsing...")
    
    for row in data_rows:
        cells = row.query_selector_all("td")
        if not cells or len(cells) < len(headers):
            continue
            
        charger_name = cells[header_map["Charger Name"]].inner_text().strip()
        charger_name = " ".join(charger_name.split())  # normalize spacing
        
        ocpp_id = cells[header_map["OCPP ID"]].inner_text().strip()
        
        # Determine status
        status = parse_status_from_cell(cells[header_map["Charger Name"]], page)
        
        if ocpp_id:
            all_chargers.append({
                "Charger Name": charger_name,
                "OCPP ID": ocpp_id,
                "Status": status
            })
            
    return all_chargers

def update_excel_vlookup_and_summary(excel_path, live_statuses):
    """
    Updates the master excel file Chargers Master.xlsx
    using OCPP ID to perform VLOOKUP status updates and recalculate pivot summaries.
    """
    log(f"Loading master Excel file {excel_path} to perform VLOOKUP and update summaries...")
    wb = openpyxl.load_workbook(excel_path)
    
    # Update 'Chargers List'
    ws_list = wb['Chargers List']
    
    # Find column indices
    headers = [cell.value for cell in next(ws_list.iter_rows(max_row=1))]
    ocpp_idx = headers.index('OCPP ID')
    
    # Find the second "Charger Status" column (containing status info)
    status_indices = [i for i, h in enumerate(headers) if h == 'Charger Status']
    status_idx = status_indices[1] if len(status_indices) > 1 else 21
    
    log(f"Matching OCPP ID in column {ocpp_idx+1} to update status in column {status_idx+1}")
    
    updated_count = 0
    for row in ws_list.iter_rows(min_row=2):
        ocpp_id = row[ocpp_idx].value
        if ocpp_id:
            ocpp_id_str = str(ocpp_id).strip()
            if ocpp_id_str in live_statuses:
                new_status = live_statuses[ocpp_id_str]
                row[status_idx].value = new_status
                updated_count += 1
            else:
                # Default to Offline if not found in live scrape
                row[status_idx].value = "Offline"
                
    log(f"VLOOKUP completed: Updated {updated_count} rows in Chargers List.")

    # Calculate summaries to update Sheet1
    data_rows = []
    for row in ws_list.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        data_rows.append({
            'EO': row[headers.index('EO')],
            'Power Rating': row[headers.index('Power Rating')],
            'Charger': row[headers.index('Charger')],
            'Status': row[status_idx]
        })
    df = pd.DataFrame(data_rows)
    
    zonal_leads = ["Bhupesh Bajiya", "Sourabh Paul", "Vasa Nagamallik"]
    ws_sheet1 = wb['Sheet1']
    
    # Left Table: All Chargers
    all_offline_total = 0
    all_online_total = 0
    
    for lead_idx, lead in enumerate(zonal_leads):
        lead_df = df[df['EO'] == lead]
        offline = len(lead_df[lead_df['Status'] == 'Offline'])
        online = len(lead_df[lead_df['Status'] == 'Online'])
        total = offline + online
        
        all_offline_total += offline
        all_online_total += online
        
        row_num = 5 + lead_idx
        ws_sheet1[f'B{row_num}'].value = offline
        ws_sheet1[f'C{row_num}'].value = online
        ws_sheet1[f'D{row_num}'].value = total
        
    # Grand Total row (Row 8)
    ws_sheet1['B8'].value = all_offline_total
    ws_sheet1['C8'].value = all_online_total
    ws_sheet1['D8'].value = all_offline_total + all_online_total
    
    # Percentage row (Row 9)
    total_all = all_offline_total + all_online_total
    ws_sheet1['B9'].value = (all_offline_total / total_all) * 100 if total_all > 0 else 0
    ws_sheet1['C9'].value = (all_online_total / total_all) * 100 if total_all > 0 else 0
    
    # Right Table: 60KW FAME Chargers Status
    fame_df_all = df[df['Charger'] == 'FAME']
    fame_offline_total = 0
    fame_online_total = 0
    
    for lead_idx, lead in enumerate(zonal_leads):
        lead_fame = fame_df_all[fame_df_all['EO'] == lead]
        offline = len(lead_fame[lead_fame['Status'] == 'Offline'])
        online = len(lead_fame[lead_fame['Status'] == 'Online'])
        total = offline + online
        
        fame_offline_total += offline
        fame_online_total += online
        
        row_num = 5 + lead_idx
        ws_sheet1[f'G{row_num}'].value = offline
        ws_sheet1[f'H{row_num}'].value = online
        ws_sheet1[f'I{row_num}'].value = total
        
    # Grand Total row (Row 8)
    ws_sheet1['G8'].value = fame_offline_total
    ws_sheet1['H8'].value = fame_online_total
    ws_sheet1['I8'].value = fame_offline_total + fame_online_total
    
    wb.save(excel_path)
    log("Master Excel file saved successfully with VLOOKUP and Sheet1 updates.")
    
    # Save a timestamped copy to Uptime folder
    now_str = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    uptime_path = os.path.join(UPTIME_DIR, f"Chargers Master_{now_str}.xlsx")
    wb.save(uptime_path)
    log(f"Hourly export saved successfully to: {uptime_path}")
    
    # Extract complete charger data list for HTML dashboard injection
    dashboard_chargers = []
    ws_list_val = wb['Chargers List']
    for r_idx, row in enumerate(ws_list_val.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:
            continue
        dashboard_chargers.append({
            "name": row[2],
            "station": row[3],
            "bl_code": row[4],
            "eo": row[5],
            "so": row[6],
            "ocpp": row[8],
            "city": row[9],
            "state": row[10],
            "power": row[17],
            "type": row[20],
            "status": row[status_idx]
        })
        
    return dashboard_chargers

def update_dashboard_html(chargers_data, excel_filename):
    if not os.path.exists(DASHBOARD_HTML_PATH):
        log(f"[WARNING] {DASHBOARD_HTML_PATH} not found. Skipping dashboard injection.")
        return
        
    with open(DASHBOARD_HTML_PATH, "r", encoding="utf-8") as f:
        html_content = f.read()
        
    payload = {
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "file_name": os.path.basename(excel_filename),
        "chargers": chargers_data
    }
    
    # Inject JSON payload into script tag
    pattern = r'(<script id="dashboard-data" type="application/json">)(.*?)(</script>)'
    replacement = f'\\1\n{json.dumps(payload, indent=4)}\n\\3'
    
    updated_html = re.sub(pattern, replacement, html_content, flags=re.DOTALL)
    
    with open(DASHBOARD_HTML_PATH, "w", encoding="utf-8") as f:
        f.write(updated_html)
        
    log("dashboard.html successfully updated with live metrics.")

def main():
    if not os.path.exists(STATE_FILE):
        log(f"[ERROR] '{STATE_FILE}' not found. Please run 'python login.py' first to authenticate.")
        sys.exit(1)
        
    log("=" * 60)
    log("BPCL EVCS Live Monitor - Service Started (Hourly)")
    log(f"Monitoring URL: {URL}")
    log(f"Check Interval: {CHECK_INTERVAL_SECONDS} seconds")
    log("=" * 60)
    
    while True:
        try:
            log("Starting status check run...")
            
            # 1. target path to master
            excel_path = EXCEL_PATH
            if not os.path.exists(excel_path):
                log(f"[ERROR] Master Excel file not found at {excel_path}!")
                time.sleep(300)
                continue
            
            # 2. Run Playwright scraper
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context(
                    storage_state=STATE_FILE,
                    viewport={"width": 1920, "height": 1080}
                )
                page = context.new_page()
                
                log(f"Navigating to {URL}...")
                try:
                    page.goto(URL, wait_until="load", timeout=60000)
                except Exception as e:
                    log(f"[WARNING] Navigation timeout: {e}")
                    
                # Check for redirection to auth login
                if "/auth" in page.url:
                    log("[ERROR] Session has expired. Please run 'python login.py' again.")
                    browser.close()
                    sys.exit(1)
                    
                time.sleep(5)
                
                # Switch dropdown view to 100 rows
                change_rows_per_page_to_100(page)
                
                # Scrape live statuses
                scraped_data = scrape_chargers(page)
                log(f"Scraped {len(scraped_data)} live chargers.")
                
                browser.close()
                
            if scraped_data:
                # Map live data into a lookup dictionary
                live_statuses = {c["OCPP ID"]: c["Status"] for c in scraped_data}
                
                # 3. Perform VLOOKUP updates on the Excel file and get dashboard list
                chargers_list = update_excel_vlookup_and_summary(excel_path, live_statuses)
                
                # 4. Inject live data to dashboard.html
                update_dashboard_html(chargers_list, excel_path)
            else:
                log("[WARNING] No live status data scraped. Skipping Excel update.")
                
        except Exception as e:
            log(f"[ERROR] Error in check run: {e}")
            traceback.print_exc()
            
        log(f"Run completed. Sleeping for {CHECK_INTERVAL_SECONDS} seconds...")
        time.sleep(CHECK_INTERVAL_SECONDS)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("Monitoring stopped by user.")
        sys.exit(0)
